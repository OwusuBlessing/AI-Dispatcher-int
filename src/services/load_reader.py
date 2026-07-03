"""Read loads from Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from configs.load_config import LoadsIngestConfig, get_loads_ingest_config


def read_load_rows(
    path: Path | str,
    *,
    config: LoadsIngestConfig | None = None,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Read raw load rows and the resolved logical-field header map."""
    ingest_config = config or get_loads_ingest_config()
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Loads file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if ingest_config.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{ingest_config.sheet_name}' not found. "
                f"Available: {workbook.sheetnames}"
            )

        worksheet = workbook[ingest_config.sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return [], {}

        headers = [_header_text(column) for column in rows[0]]
        headers = _trim_trailing_empty_headers(headers)
        header_map = ingest_config.resolve_header_map(headers)
        ingest_config.require_mapped_fields(header_map)

        id_header = header_map[ingest_config.id_field]
        raw_rows: list[dict[str, Any]] = []
        for index, row in enumerate(rows[1:], start=2):
            row_dict = {
                headers[i]: row[i] if i < len(row) else None
                for i in range(len(headers))
            }
            if _is_missing(row_dict.get(id_header), ingest_config):
                continue
            row_dict["_source_row"] = index
            row_dict["_header_map"] = header_map
            raw_rows.append(row_dict)

        return raw_rows, header_map
    finally:
        workbook.close()


def _header_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _trim_trailing_empty_headers(headers: list[str]) -> list[str]:
    trimmed = list(headers)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _is_missing(value: Any, config: LoadsIngestConfig) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text.casefold() in {item.casefold() for item in config.missing_values}
